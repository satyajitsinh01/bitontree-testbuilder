"""Bulk candidate import from CSV/XLSX (research R7). Per-row validation with a
row-level error report; valid rows import even when others fail (FR-012)."""

import csv
import io
import math
import re

from openpyxl import load_workbook

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
# Indian mobile: exactly 10 digits starting 6-9, optional +91 prefix
INDIAN_MOBILE_RE = re.compile(r"^(?:\+91)?[6-9]\d{9}$")
COLUMNS = ["studentId", "name", "email", "phone", "cgpa"]


def normalize_phone(raw: str) -> str:
    return re.sub(r"[\s()\-]", "", raw)


def parse_rows(filename: str, content: bytes) -> list[dict]:
    """Returns raw dict rows keyed by COLUMNS."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(io.BytesIO(content), read_only=True)
        sheet = workbook.active
        rows = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                header = ["" if value is None else str(value).strip() for value in row]
                if header[: len(COLUMNS)] != COLUMNS:
                    raise ValueError(f"expected columns: {','.join(COLUMNS)}")
                continue
            values = ["" if v is None else str(v) for v in row[: len(COLUMNS)]]
            values += [""] * (len(COLUMNS) - len(values))
            rows.append(dict(zip(COLUMNS, values, strict=True)))
        return rows
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    header = next(reader, [])
    if [value.strip() for value in header] != COLUMNS:
        raise ValueError(f"expected columns: {','.join(COLUMNS)}")
    rows = []
    for row in reader:
        values = [v.strip() for v in row[: len(COLUMNS)]]
        values += [""] * (len(COLUMNS) - len(values))
        rows.append(dict(zip(COLUMNS, values, strict=True)))
    return rows


def validate_row(row: dict) -> tuple[dict | None, str | None]:
    """Validate and normalize one candidate import row."""
    student_id = row.get("studentId", "").strip()
    name = row.get("name", "").strip()
    email = row.get("email", "").strip().lower()
    phone = normalize_phone(row.get("phone", "").strip())
    if not student_id:
        return None, "studentId is required"
    if not name:
        return None, "name is required"
    if not EMAIL_RE.match(email):
        return None, f"invalid email: {email or '(empty)'}"
    if phone and not INDIAN_MOBILE_RE.match(phone):
        return None, f"invalid phone (expected 10-digit Indian mobile): {phone}"
    try:
        cgpa = float(row.get("cgpa", "").strip())
    except (TypeError, ValueError):
        return None, "cgpa must be a number between 0 and 10"
    if not math.isfinite(cgpa) or not 0 <= cgpa <= 10:
        return None, "cgpa must be a number between 0 and 10"
    return (
        {
            "student_id": student_id,
            "name": name,
            "email": email,
            "phone": phone,
            "cgpa": cgpa,
        },
        None,
    )


CSV_TEMPLATE = (
    "studentId,name,email,phone,cgpa\n"
    "STU-1001,Jane Doe,jane@example.com,+919876543210,8.75\n"
)
